"""
Statistics Routes
Maneja los endpoints de estadísticas del sistema
Sigue el patrón de rutas usado en el proyecto
"""
from flask import Blueprint, request, jsonify, send_file
from services.statistics_service import StatisticsService
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

statistics_bp = Blueprint("statistics", __name__)


# ============================================================================
# ENDPOINTS DE ESTADÍSTICAS
# ============================================================================

@statistics_bp.route("/dashboard", methods=["GET"])
def get_dashboard_overview():
    """
    Obtiene resumen general del dashboard
    GET /statistics/dashboard
    
    Returns:
        JSON con métricas principales del sistema
    """
    try:
        overview = StatisticsService.get_dashboard_overview()
        return jsonify({
            "success": True,
            "data": overview
        }), 200
        
    except Exception as e:
        print(f"Error en /statistics/dashboard: {str(e)}")
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


@statistics_bp.route("/temporal", methods=["GET"])
def get_temporal_analysis():
    """
    Obtiene análisis temporal de casos
    GET /statistics/temporal?period=month&days=30
    
    Query Parameters:
        period: Tipo de periodo (day, week, month)
        days: Número de días hacia atrás
        
    Returns:
        JSON con análisis temporal
    """
    try:
        period = request.args.get("period", "month")
        days = request.args.get("days", type=int)
        
        temporal_data = StatisticsService.get_temporal_analysis(period, days)
        return jsonify({
            "success": True,
            "data": temporal_data
        }), 200
        
    except Exception as e:
        print(f"Error en /statistics/temporal: {str(e)}")
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


@statistics_bp.route("/detection-metrics", methods=["GET"])
def get_detection_metrics():
    """
    Obtiene métricas de detección facial
    GET /statistics/detection-metrics
    
    Returns:
        JSON con métricas de detección y falsos positivos
    """
    try:
        metrics = StatisticsService.get_detection_metrics()
        return jsonify({
            "success": True,
            "data": metrics
        }), 200
        
    except Exception as e:
        print(f"Error en /statistics/detection-metrics: {str(e)}")
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


@statistics_bp.route("/heatmap", methods=["GET"])
def get_heatmap_data():
    """
    Obtiene datos para mapa de calor de detecciones
    GET /statistics/heatmap
    
    Returns:
        JSON con datos de ubicaciones y conteos
    """
    try:
        heatmap_data = StatisticsService.get_heatmap_data()
        return jsonify({
            "success": True,
            "data": heatmap_data
        }), 200
        
    except Exception as e:
        print(f"Error en /statistics/heatmap: {str(e)}")
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


@statistics_bp.route("/demographics", methods=["GET"])
def get_demographic_analysis():
    """
    Obtiene análisis demográfico de casos
    GET /statistics/demographics
    
    Returns:
        JSON con distribución por edad y otros datos demográficos
    """
    try:
        demographics = StatisticsService.get_demographic_analysis()
        return jsonify({
            "success": True,
            "data": demographics
        }), 200
        
    except Exception as e:
        print(f"Error en /statistics/demographics: {str(e)}")
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


@statistics_bp.route("/cameras", methods=["GET"])
def get_camera_statistics():
    """
    Obtiene estadísticas por cámara
    GET /statistics/cameras
    
    Returns:
        JSON con estadísticas de cada cámara
    """
    try:
        camera_stats = StatisticsService.get_camera_statistics()
        return jsonify({
            "success": True,
            "data": camera_stats
        }), 200
        
    except Exception as e:
        print(f"Error en /statistics/cameras: {str(e)}")
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


@statistics_bp.route("/performance", methods=["GET"])
def get_performance_metrics():
    """
    Obtiene métricas de rendimiento del sistema
    GET /statistics/performance
    
    Returns:
        JSON con métricas de rendimiento y eficiencia
    """
    try:
        performance = StatisticsService.get_performance_metrics()
        return jsonify({
            "success": True,
            "data": performance
        }), 200
        
    except Exception as e:
        print(f"Error en /statistics/performance: {str(e)}")
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


@statistics_bp.route("/report/complete", methods=["GET"])
def get_complete_report():
    """
    Genera un reporte completo del sistema
    GET /statistics/report/complete
    
    Returns:
        JSON con todas las métricas disponibles
    """
    try:
        report = StatisticsService.get_complete_report()
        return jsonify({
            "success": True,
            "data": report
        }), 200
        
    except Exception as e:
        print(f"Error en /statistics/report/complete: {str(e)}")
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


# ============================================================================
# ENDPOINTS DE EXPORTACIÓN
# ============================================================================

@statistics_bp.route("/export/pdf", methods=["POST"])
def export_report_pdf():
    """
    Exporta reporte a PDF
    POST /statistics/export/pdf
    
    Body:
        report_type: Tipo de reporte (dashboard, complete, custom)
        filters: Filtros opcionales
        
    Returns:
        PDF file
    """
    try:
        data = request.get_json() or {}
        report_type = data.get("report_type", "complete")
        
        # Obtener datos
        overview = StatisticsService.get_dashboard_overview()
        
        # Crear PDF en memoria
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []
        
        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#2F5496'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#2F5496'),
            spaceAfter=12,
            spaceBefore=12
        )
        
        # Título principal
        title = Paragraph("REPORTE DE ESTADÍSTICAS - FACEFIND", title_style)
        elements.append(title)
        
        subtitle = Paragraph(
            f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            styles['Normal']
        )
        elements.append(subtitle)
        elements.append(Spacer(1, 0.3*inch))
        
        # Resumen General
        elements.append(Paragraph("RESUMEN GENERAL", heading_style))
        
        summary = overview.get('summary', {})
        summary_data = [
            ['Métrica', 'Valor'],
            ['Casos Totales', str(summary.get('total_cases', 0))],
            ['Casos Activos', str(summary.get('active_cases', 0))],
            ['Casos Pendientes', str(summary.get('pending_cases', 0))],
            ['Casos Resueltos', str(summary.get('resolved_cases', 0))],
            ['Usuarios Activos', str(summary.get('active_users', 0))],
            ['Detecciones Totales', str(summary.get('total_detections', 0))],
            ['Tasa de Resolución', f"{summary.get('resolution_rate', 0)}%"]
        ]
        
        summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 0.3*inch))
        
        # Distribución por Estado
        elements.append(Paragraph("DISTRIBUCIÓN POR ESTADO", heading_style))
        
        status_data = [['Estado', 'Cantidad', 'Porcentaje']]
        for item in overview.get('status_distribution', []):
            status_data.append([
                item.get('status', 'N/A'),
                str(item.get('count', 0)),
                f"{item.get('percentage', 0):.1f}%"
            ])
        
        if len(status_data) > 1:
            status_table = Table(status_data, colWidths=[2*inch, 1.5*inch, 1.5*inch])
            status_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.lightblue),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
            ]))
            elements.append(status_table)
            elements.append(Spacer(1, 0.3*inch))
        
        # Demografía
        elements.append(Paragraph("DISTRIBUCIÓN DEMOGRÁFICA", heading_style))
        
        demographics = overview.get('demographics', {})
        demo_data = [['Grupo de Edad', 'Cantidad', 'Porcentaje']]
        for item in demographics.get('age_distribution', []):
            demo_data.append([
                item.get('age_group', 'N/A'),
                str(item.get('count', 0)),
                f"{item.get('percentage', 0):.1f}%"
            ])
        
        if len(demo_data) > 1:
            demo_table = Table(demo_data, colWidths=[2*inch, 1.5*inch, 1.5*inch])
            demo_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.lightgreen),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
            ]))
            elements.append(demo_table)
            elements.append(Spacer(1, 0.3*inch))
        
        # Nueva página para métricas de detección
        elements.append(PageBreak())
        
        # Métricas de Detección
        elements.append(Paragraph("MÉTRICAS DE DETECCIÓN FACIAL", heading_style))
        
        detection = overview.get('detection_metrics', {})
        detection_data = [
            ['Métrica', 'Valor'],
            ['Detecciones Totales', str(detection.get('total_detections', 0))],
            ['Coincidencias Exitosas', str(detection.get('successful_matches', 0))],
            ['Falsos Positivos', str(detection.get('false_positives', 0))],
            ['Tasa de Precisión', f"{detection.get('accuracy_rate', 0):.2f}%"],
            ['Tiempo Promedio', f"{detection.get('avg_detection_time', 0):.2f}s"]
        ]
        
        detection_table = Table(detection_data, colWidths=[3*inch, 2*inch])
        detection_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.lightyellow),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ]))
        elements.append(detection_table)
        
        # Construir PDF
        doc.build(elements)
        buffer.seek(0)
        
        filename = f"reporte_facefind_{report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"Error en /statistics/export/pdf: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


@statistics_bp.route("/export/excel", methods=["POST"])
def export_report_excel():
    """
    Exporta reporte a Excel
    POST /statistics/export/excel
    
    Body:
        report_type: Tipo de reporte (dashboard, complete, custom)
        filters: Filtros opcionales
        
    Returns:
        Excel file
    """
    try:
        data = request.get_json() or {}
        report_type = data.get("report_type", "complete")
        
        # Obtener todos los datos necesarios
        overview = StatisticsService.get_dashboard_overview()
        
        # Crear workbook
        wb = Workbook()
        
        # Estilos
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, size=14, color="2F5496")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Hoja 1: Resumen General
        ws1 = wb.active
        ws1.title = "Resumen General"
        
        # Título
        ws1['A1'] = 'REPORTE DE ESTADÍSTICAS - FACEFIND'
        ws1['A1'].font = title_font
        ws1['A2'] = f'Generado: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        
        # Resumen
        ws1['A4'] = 'Métrica'
        ws1['B4'] = 'Valor'
        ws1['A4'].font = header_font
        ws1['B4'].font = header_font
        ws1['A4'].fill = header_fill
        ws1['B4'].fill = header_fill
        
        summary = overview.get('summary', {})
        metrics = [
            ('Casos Totales', summary.get('total_cases', 0)),
            ('Casos Activos', summary.get('active_cases', 0)),
            ('Casos Pendientes', summary.get('pending_cases', 0)),
            ('Casos Resueltos', summary.get('resolved_cases', 0)),
            ('Usuarios Activos', summary.get('active_users', 0)),
            ('Detecciones Totales', summary.get('total_detections', 0)),
            ('Tasa de Resolución', f"{summary.get('resolution_rate', 0)}%")
        ]
        
        row = 5
        for metric, value in metrics:
            ws1[f'A{row}'] = metric
            ws1[f'B{row}'] = value
            ws1[f'A{row}'].border = border
            ws1[f'B{row}'].border = border
            row += 1
        
        # Ajustar ancho de columnas
        ws1.column_dimensions['A'].width = 25
        ws1.column_dimensions['B'].width = 20
        
        # Hoja 2: Casos por Estado
        ws2 = wb.create_sheet(title="Casos por Estado")
        ws2['A1'] = 'DISTRIBUCIÓN DE CASOS POR ESTADO'
        ws2['A1'].font = title_font
        
        ws2['A3'] = 'Estado'
        ws2['B3'] = 'Cantidad'
        ws2['C3'] = 'Porcentaje'
        for cell in ['A3', 'B3', 'C3']:
            ws2[cell].font = header_font
            ws2[cell].fill = header_fill
            ws2[cell].border = border
        
        status_dist = overview.get('status_distribution', [])
        row = 4
        for item in status_dist:
            ws2[f'A{row}'] = item.get('status', 'N/A')
            ws2[f'B{row}'] = item.get('count', 0)
            ws2[f'C{row}'] = f"{item.get('percentage', 0):.1f}%"
            for col in ['A', 'B', 'C']:
                ws2[f'{col}{row}'].border = border
            row += 1
        
        ws2.column_dimensions['A'].width = 20
        ws2.column_dimensions['B'].width = 15
        ws2.column_dimensions['C'].width = 15
        
        # Hoja 3: Demografía
        ws3 = wb.create_sheet(title="Demografía")
        ws3['A1'] = 'DISTRIBUCIÓN DEMOGRÁFICA'
        ws3['A1'].font = title_font
        
        ws3['A3'] = 'Grupo de Edad'
        ws3['B3'] = 'Cantidad'
        ws3['C3'] = 'Porcentaje'
        for cell in ['A3', 'B3', 'C3']:
            ws3[cell].font = header_font
            ws3[cell].fill = header_fill
            ws3[cell].border = border
        
        demographics = overview.get('demographics', {})
        age_dist = demographics.get('age_distribution', [])
        row = 4
        for item in age_dist:
            ws3[f'A{row}'] = item.get('age_group', 'N/A')
            ws3[f'B{row}'] = item.get('count', 0)
            ws3[f'C{row}'] = f"{item.get('percentage', 0):.1f}%"
            for col in ['A', 'B', 'C']:
                ws3[f'{col}{row}'].border = border
            row += 1
        
        ws3.column_dimensions['A'].width = 20
        ws3.column_dimensions['B'].width = 15
        ws3.column_dimensions['C'].width = 15
        
        # Hoja 4: Métricas de Detección
        ws4 = wb.create_sheet(title="Métricas Detección")
        ws4['A1'] = 'MÉTRICAS DE DETECCIÓN FACIAL'
        ws4['A1'].font = title_font
        
        detection = overview.get('detection_metrics', {})
        ws4['A3'] = 'Métrica'
        ws4['B3'] = 'Valor'
        ws4['A3'].font = header_font
        ws4['B3'].font = header_font
        ws4['A3'].fill = header_fill
        ws4['B3'].fill = header_fill
        
        detection_metrics = [
            ('Detecciones Totales', detection.get('total_detections', 0)),
            ('Coincidencias Exitosas', detection.get('successful_matches', 0)),
            ('Falsos Positivos', detection.get('false_positives', 0)),
            ('Tasa de Precisión', f"{detection.get('accuracy_rate', 0):.2f}%"),
            ('Tiempo Promedio', f"{detection.get('avg_detection_time', 0):.2f}s")
        ]
        
        row = 4
        for metric, value in detection_metrics:
            ws4[f'A{row}'] = metric
            ws4[f'B{row}'] = value
            ws4[f'A{row}'].border = border
            ws4[f'B{row}'].border = border
            row += 1
        
        ws4.column_dimensions['A'].width = 25
        ws4.column_dimensions['B'].width = 20
        
        # Guardar en BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"reporte_facefind_{report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"Error en /statistics/export/excel: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500
