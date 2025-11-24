"""
Report Service - Servicio para exportación de reportes
Genera reportes en Excel y CSV con filtros personalizados
"""
from typing import Optional, Dict, List
from datetime import datetime
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.drawing.image import Image
import csv
from services.supabase_client import supabase


class ReportService:
    """Servicio para generación de reportes con filtros"""

    @staticmethod
    def generar_reporte_excel(
        fecha_inicio: Optional[datetime] = None,
        fecha_fin: Optional[datetime] = None,
        estado: Optional[str] = None,
        usuario_id: Optional[int] = None,
        camara_id: Optional[int] = None
    ) -> BytesIO:
        """
        Genera un reporte detallado en Excel con gráficos
        
        Args:
            fecha_inicio: Fecha de inicio del rango
            fecha_fin: Fecha de fin del rango
            estado: Filtro por estado de alerta
            usuario_id: Filtro por usuario
            camara_id: Filtro por cámara
            
        Returns:
            BytesIO con el archivo Excel
        """
        # Obtener datos filtrados
        alertas_data = ReportService._obtener_alertas_filtradas(
            fecha_inicio, fecha_fin, estado, usuario_id, camara_id
        )
        
        # Crear libro de Excel
        wb = openpyxl.Workbook()
        
        # Hoja 1: Datos de alertas
        ws_alertas = wb.active
        ws_alertas.title = "Alertas"
        ReportService._crear_hoja_alertas(ws_alertas, alertas_data)
        
        # Hoja 2: Resumen estadístico
        ws_stats = wb.create_sheet("Estadísticas")
        ReportService._crear_hoja_estadisticas(ws_stats, alertas_data)
        
        # Hoja 3: Gráficos
        ws_charts = wb.create_sheet("Gráficos")
        ReportService._crear_hoja_graficos(ws_charts, alertas_data)
        
        # Agregar marca de agua institucional en todas las hojas
        ReportService._agregar_marca_agua(wb)
        
        # Guardar en BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output

    @staticmethod
    def generar_reporte_csv(
        fecha_inicio: Optional[datetime] = None,
        fecha_fin: Optional[datetime] = None,
        estado: Optional[str] = None,
        usuario_id: Optional[int] = None,
        camara_id: Optional[int] = None
    ) -> BytesIO:
        """
        Genera un reporte en formato CSV
        
        Args:
            fecha_inicio: Fecha de inicio del rango
            fecha_fin: Fecha de fin del rango
            estado: Filtro por estado de alerta
            usuario_id: Filtro por usuario
            camara_id: Filtro por cámara
            
        Returns:
            BytesIO con el archivo CSV
        """
        # Obtener datos filtrados
        alertas_data = ReportService._obtener_alertas_filtradas(
            fecha_inicio, fecha_fin, estado, usuario_id, camara_id
        )
        
        # Crear CSV
        output = BytesIO()
        
        # Escribir con codificación UTF-8
        import io
        text_output = io.StringIO()
        writer = csv.writer(text_output)
        
        # Encabezados
        writer.writerow([
            'ID Alerta',
            'Fecha/Hora',
            'Persona Desaparecida',
            'Ubicación',
            'Similitud (%)',
            'Estado',
            'Prioridad',
            'Cámara',
            'Usuario Reportante',
            'Latitud',
            'Longitud',
            'Falso Positivo'
        ])
        
        # Datos
        for alerta in alertas_data:
            writer.writerow([
                alerta['id'],
                alerta['timestamp'],
                alerta['persona_nombre'],
                alerta['ubicacion'],
                f"{alerta['similitud'] * 100:.1f}",
                alerta['estado'],
                alerta['prioridad'],
                alerta['camara_info'],
                alerta['usuario_nombre'],
                alerta['latitud'],
                alerta['longitud'],
                'Sí' if alerta['falso_positivo'] else 'No'
            ])
        
        # Convertir a bytes
        output.write(text_output.getvalue().encode('utf-8-sig'))  # BOM para Excel
        output.seek(0)
        
        return output

    @staticmethod
    def _obtener_alertas_filtradas(
        fecha_inicio: Optional[datetime],
        fecha_fin: Optional[datetime],
        estado: Optional[str],
        usuario_id: Optional[int],
        camara_id: Optional[int]
    ) -> List[Dict]:
        """
        Obtiene alertas con filtros aplicados y datos enriquecidos
        
        Returns:
            Lista de diccionarios con datos de alertas
        """
        try:
            # Query con JOINs para obtener todos los datos necesarios
            query = supabase.table("Alerta")\
                .select("""
                    *,
                    Camara(id, ubicacion, ip, type),
                    Caso(
                        id,
                        usuario_id,
                        status,
                        PersonaDesaparecida(nombre_completo),
                        Usuario(nombre, email)
                    )
                """)
            
            # Aplicar filtros
            if fecha_inicio:
                query = query.gte("timestamp", fecha_inicio.isoformat())
            if fecha_fin:
                query = query.lte("timestamp", fecha_fin.isoformat())
            if estado:
                query = query.eq("estado", estado.upper())
            if camara_id:
                query = query.eq("camara_id", camara_id)
            
            response = query.order("timestamp", desc=True).execute()
            
            if not response.data:
                return []
            
            # Procesar y enriquecer datos
            alertas_procesadas = []
            for alerta in response.data:
                # Filtro adicional por usuario (a nivel de caso)
                if usuario_id:
                    if not alerta.get('Caso') or alerta['Caso'].get('usuario_id') != usuario_id:
                        continue
                
                # Extraer datos anidados
                caso = alerta.get('Caso', {})
                persona = caso.get('PersonaDesaparecida', {})
                usuario = caso.get('Usuario', {})
                camara = alerta.get('Camara', {})
                
                alertas_procesadas.append({
                    'id': alerta['id'],
                    'timestamp': alerta['timestamp'],
                    'similitud': alerta.get('similitud', 0),
                    'estado': alerta.get('estado', 'PENDIENTE'),
                    'prioridad': alerta.get('prioridad', 'MEDIA'),
                    'ubicacion': alerta.get('ubicacion', 'N/A'),
                    'latitud': alerta.get('latitud'),
                    'longitud': alerta.get('longitud'),
                    'falso_positivo': alerta.get('falso_positivo', False),
                    'persona_nombre': persona.get('nombre_completo', 'Desconocido'),
                    'usuario_nombre': usuario.get('nombre', 'N/A'),
                    'usuario_email': usuario.get('email', 'N/A'),
                    'camara_info': f"{camara.get('type', 'N/A')} - {camara.get('ubicacion', 'N/A')}",
                    'caso_id': caso.get('id'),
                    'caso_status': caso.get('status', 'N/A')
                })
            
            return alertas_procesadas
            
        except Exception as e:
            print(f"Error obteniendo alertas filtradas: {e}")
            import traceback
            traceback.print_exc()
            return []

    @staticmethod
    def _crear_hoja_alertas(ws, alertas_data: List[Dict]):
        """Crea la hoja de alertas con formato"""
        # Título
        ws.merge_cells('A1:L1')
        title_cell = ws['A1']
        title_cell.value = "REPORTE DE ALERTAS - FACEFIND"
        title_cell.font = Font(size=16, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30
        
        # Información del reporte
        ws['A2'] = f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A3'] = f"Total de alertas: {len(alertas_data)}"
        
        # Encabezados
        headers = [
            'ID', 'Fecha/Hora', 'Persona', 'Ubicación', 'Similitud %',
            'Estado', 'Prioridad', 'Cámara', 'Usuario', 'Lat', 'Lon', 'Falso Positivo'
        ]
        
        header_fill = PatternFill(start_color="764BA2", end_color="764BA2", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Datos
        for row_idx, alerta in enumerate(alertas_data, 6):
            ws.cell(row=row_idx, column=1, value=alerta['id'])
            ws.cell(row=row_idx, column=2, value=alerta['timestamp'])
            ws.cell(row=row_idx, column=3, value=alerta['persona_nombre'])
            ws.cell(row=row_idx, column=4, value=alerta['ubicacion'])
            ws.cell(row=row_idx, column=5, value=f"{alerta['similitud'] * 100:.1f}")
            ws.cell(row=row_idx, column=6, value=alerta['estado'])
            ws.cell(row=row_idx, column=7, value=alerta['prioridad'])
            ws.cell(row=row_idx, column=8, value=alerta['camara_info'])
            ws.cell(row=row_idx, column=9, value=alerta['usuario_nombre'])
            ws.cell(row=row_idx, column=10, value=alerta['latitud'])
            ws.cell(row=row_idx, column=11, value=alerta['longitud'])
            ws.cell(row=row_idx, column=12, value='Sí' if alerta['falso_positivo'] else 'No')
            
            # Colorear según prioridad
            if alerta['prioridad'] == 'ALTA':
                for col in range(1, 13):
                    ws.cell(row=row_idx, column=col).fill = PatternFill(
                        start_color="FFE6E6", end_color="FFE6E6", fill_type="solid"
                    )
        
        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 30
        ws.column_dimensions['I'].width = 20
        ws.column_dimensions['J'].width = 12
        ws.column_dimensions['K'].width = 12
        ws.column_dimensions['L'].width = 15

    @staticmethod
    def _crear_hoja_estadisticas(ws, alertas_data: List[Dict]):
        """Crea la hoja de estadísticas"""
        # Título
        ws.merge_cells('A1:D1')
        title_cell = ws['A1']
        title_cell.value = "ESTADÍSTICAS"
        title_cell.font = Font(size=14, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center")
        
        # Resumen general
        ws['A3'] = "RESUMEN GENERAL"
        ws['A3'].font = Font(bold=True, size=12)
        
        ws['A4'] = "Total de Alertas:"
        ws['B4'] = len(alertas_data)
        
        # Por estado
        estados = {}
        for alerta in alertas_data:
            estado = alerta['estado']
            estados[estado] = estados.get(estado, 0) + 1
        
        row = 6
        ws[f'A{row}'] = "POR ESTADO"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        for estado, count in estados.items():
            ws[f'A{row}'] = estado
            ws[f'B{row}'] = count
            ws[f'C{row}'] = f"{count/len(alertas_data)*100:.1f}%" if alertas_data else "0%"
            row += 1
        
        # Por prioridad
        row += 1
        ws[f'A{row}'] = "POR PRIORIDAD"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        prioridades = {}
        for alerta in alertas_data:
            prioridad = alerta['prioridad']
            prioridades[prioridad] = prioridades.get(prioridad, 0) + 1
        
        for prioridad, count in prioridades.items():
            ws[f'A{row}'] = prioridad
            ws[f'B{row}'] = count
            ws[f'C{row}'] = f"{count/len(alertas_data)*100:.1f}%" if alertas_data else "0%"
            row += 1
        
        # Similitud promedio
        if alertas_data:
            similitud_promedio = sum(a['similitud'] for a in alertas_data) / len(alertas_data)
            row += 1
            ws[f'A{row}'] = "Similitud Promedio:"
            ws[f'B{row}'] = f"{similitud_promedio * 100:.1f}%"
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15

    @staticmethod
    def _crear_hoja_graficos(ws, alertas_data: List[Dict]):
        """Crea la hoja de gráficos"""
        if not alertas_data:
            ws['A1'] = "No hay datos para generar gráficos"
            return
        
        # Título
        ws['A1'] = "GRÁFICOS ESTADÍSTICOS"
        ws['A1'].font = Font(size=14, bold=True)
        
        # Preparar datos para gráfico de estados
        estados = {}
        for alerta in alertas_data:
            estado = alerta['estado']
            estados[estado] = estados.get(estado, 0) + 1
        
        # Datos para gráfico
        row = 3
        ws[f'A{row}'] = "Estado"
        ws[f'B{row}'] = "Cantidad"
        row += 1
        
        for estado, count in estados.items():
            ws[f'A{row}'] = estado
            ws[f'B{row}'] = count
            row += 1
        
        # Crear gráfico de barras
        chart = BarChart()
        chart.title = "Alertas por Estado"
        chart.x_axis.title = "Estado"
        chart.y_axis.title = "Cantidad"
        
        data = Reference(ws, min_col=2, min_row=3, max_row=row-1)
        cats = Reference(ws, min_col=1, min_row=4, max_row=row-1)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        ws.add_chart(chart, "D3")
        
        # Gráfico de pastel para prioridades
        prioridades = {}
        for alerta in alertas_data:
            prioridad = alerta['prioridad']
            prioridades[prioridad] = prioridades.get(prioridad, 0) + 1
        
        row += 2
        ws[f'A{row}'] = "Prioridad"
        ws[f'B{row}'] = "Cantidad"
        row += 1
        start_row = row
        
        for prioridad, count in prioridades.items():
            ws[f'A{row}'] = prioridad
            ws[f'B{row}'] = count
            row += 1
        
        # Gráfico de pastel
        pie = PieChart()
        pie.title = "Distribución por Prioridad"
        
        data = Reference(ws, min_col=2, min_row=start_row-1, max_row=row-1)
        cats = Reference(ws, min_col=1, min_row=start_row, max_row=row-1)
        
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(cats)
        
        ws.add_chart(pie, "D20")

    @staticmethod
    def _agregar_marca_agua(wb):
        """Agrega marca de agua institucional a todas las hojas"""
        for ws in wb.worksheets:
            # Agregar footer con marca institucional
            ws.oddFooter.center.text = "FaceFind - Sistema de Reconocimiento Facial | Generado: " + datetime.now().strftime('%Y-%m-%d %H:%M')
            ws.oddFooter.center.size = 9
